"""
Auto-indexing signals for RAG system.
Supports PDF, TXT, MD, DOCX — indexes into ChromaDB on document save.

Documents are queued and processed ONE AT A TIME by a single background
worker thread — no concurrent indexing, no ChromaDB write collisions,
no interleaved console output.
"""

import os
import queue
import threading
import logging

from django.db.models.signals import post_save
from django.dispatch import receiver

logger = logging.getLogger(__name__)

# All file types the indexer can handle
SUPPORTED_EXTENSIONS = {'.pdf', '.txt', '.md', '.text', '.docx', '.doc', '.xls', '.xlsx'}

# ── Single-worker indexing queue ─────────────────────────────────────────────
# All save signals drop a document_id into this queue.
# One daemon worker thread drains it sequentially — zero concurrency issues.
_index_queue  = queue.Queue()
_worker_lock  = threading.Lock()
_worker_alive = False


def _ensure_worker():
    """Start the background worker thread if not already running."""
    global _worker_alive
    if _worker_alive:
        return
    with _worker_lock:
        if _worker_alive:          # re-check inside lock
            return
        t = threading.Thread(
            target=_queue_worker,
            daemon=True,
            name='rag-index-worker',
        )
        t.start()
        _worker_alive = True


def _queue_worker():
    """
    Single daemon thread — drains the index queue one document at a time.
    Never spawns sub-threads; all indexing is fully sequential.
    """
    while True:
        document_id = _index_queue.get()
        try:
            _index_in_background(document_id)
        except Exception as exc:
            logger.error(
                f"[RAG] Queue worker unhandled error for doc {document_id}: {exc}",
                exc_info=True,
            )
        finally:
            _index_queue.task_done()


# ── Signal handler ────────────────────────────────────────────────────────────

@receiver(post_save, sender='documents.Document')
def auto_index_document(sender, instance, created, **kwargs):
    """
    Queue a document for indexing whenever it is saved.
    Never blocks the HTTP request — the worker picks it up asynchronously.
    """
    if not instance.file:
        return

    try:
        file_path = instance.file.path
    except Exception:
        return

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        logger.debug(f"[RAG] Skipping unsupported file type: {ext}")
        return

    # Fast pre-check — skip if already successfully indexed
    try:
        from .models import DocumentEmbedding
        emb = DocumentEmbedding.objects.get(document=instance)
        if emb.is_indexed:
            return
    except Exception:
        pass  # No record yet — worker will create one

    _ensure_worker()
    _index_queue.put(instance.id)
    logger.debug(
        f"[RAG] Doc {instance.id} queued for indexing "
        f"(queue depth: {_index_queue.qsize()})"
    )


# ── File loaders ──────────────────────────────────────────────────────────────

def _load_text_file(file_path: str, document):
    """
    Load a plain-text / markdown file as a list of LangChain Document objects.
    Each ~500-word block becomes one document to give the chunker good material.
    """
    from langchain_core.documents import Document as LCDoc

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as fh:
        raw = fh.read()

    if not raw.strip():
        return []

    paragraphs = [p.strip() for p in raw.split('\n\n') if p.strip()]
    pages, current, word_count = [], [], 0
    for para in paragraphs:
        words = len(para.split())
        if word_count + words > 500 and current:
            pages.append('\n\n'.join(current))
            current, word_count = [], 0
        current.append(para)
        word_count += words
    if current:
        pages.append('\n\n'.join(current))

    docs = []
    for i, page_text in enumerate(pages, 1):
        docs.append(LCDoc(
            page_content=page_text,
            metadata={
                'source':       file_path,
                'title':        getattr(document, 'title', os.path.basename(file_path)),
                'page':         i,
                'content_type': 'text',
                'file_type':    os.path.splitext(file_path)[1].lower().lstrip('.'),
            }
        ))
    return docs


def _load_docx_file(file_path: str, document):
    """Load a .docx file as LangChain Document objects (paragraph-by-paragraph)."""
    try:
        from docx import Document as DocxDoc
        from langchain_core.documents import Document as LCDoc
    except ImportError:
        logger.warning(
            "[RAG] python-docx not installed — cannot index .docx files. "
            "Run: pip install python-docx"
        )
        return []

    docx = DocxDoc(file_path)
    paragraphs = [p.text.strip() for p in docx.paragraphs if p.text.strip()]
    if not paragraphs:
        return []

    pages, current, word_count = [], [], 0
    for para in paragraphs:
        words = len(para.split())
        if word_count + words > 500 and current:
            pages.append('\n\n'.join(current))
            current, word_count = [], 0
        current.append(para)
        word_count += words
    if current:
        pages.append('\n\n'.join(current))

    docs = []
    for i, page_text in enumerate(pages, 1):
        docs.append(LCDoc(
            page_content=page_text,
            metadata={
                'source':       file_path,
                'title':        getattr(document, 'title', os.path.basename(file_path)),
                'page':         i,
                'content_type': 'text',
                'file_type':    'docx',
            }
        ))
    return docs


def _load_excel_file(file_path: str, document):
    """
    Load an .xls/.xlsx file as LangChain Document objects.

    Each data row is formatted as "Header: Value | Header: Value ..." so that
    every chunk is fully self-describing — the LLM always knows which column
    a value belongs to, enabling accurate date filtering and aggregations.
    """
    from langchain_core.documents import Document as LCDoc

    ext = os.path.splitext(file_path)[1].lower()
    # sheets_data: list of (sheet_name, headers, list_of_row_dicts)
    sheets_data = []

    if ext == '.xlsx':
        try:
            import openpyxl
        except ImportError:
            logger.warning(
                "[RAG] openpyxl not installed — cannot index .xlsx files. "
                "Run: pip install openpyxl"
            )
            return []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue
            headers = [str(c).strip() if c is not None else f'Col{i}' for i, c in enumerate(all_rows[0])]
            rows = []
            for row in all_rows[1:]:
                cells = [str(c).strip() if c is not None else '' for c in row]
                if any(c for c in cells):
                    rows.append(cells)
            if rows:
                sheets_data.append((sheet_name, headers, rows))
        wb.close()

    else:  # .xls
        try:
            import xlrd
        except ImportError:
            logger.warning(
                "[RAG] xlrd not installed — cannot index .xls files. "
                "Run: pip install xlrd"
            )
            return []
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
            if sheet.nrows < 2:
                continue
            headers = [str(sheet.cell(0, cx).value).strip() or f'Col{cx}' for cx in range(sheet.ncols)]
            rows = []
            for rx in range(1, sheet.nrows):
                cells = [str(sheet.cell(rx, cx).value).strip() for cx in range(sheet.ncols)]
                if any(c for c in cells):
                    rows.append(cells)
            if rows:
                sheets_data.append((sheet.name, headers, rows))

    if not sheets_data:
        return []

    docs = []
    for sheet_name, headers, rows in sheets_data:
        # Format each row as "Header: Value | Header: Value ..."
        # Group ~50 rows per chunk so each chunk fits in context comfortably
        ROWS_PER_CHUNK = 30
        for chunk_start in range(0, len(rows), ROWS_PER_CHUNK):
            chunk_rows = rows[chunk_start:chunk_start + ROWS_PER_CHUNK]
            lines = []
            for row in chunk_rows:
                pairs = ' | '.join(
                    f'{h}: {v}' for h, v in zip(headers, row) if v
                )
                if pairs:
                    lines.append(pairs)
            if not lines:
                continue
            page_text = '\n'.join(lines)
            page_num = (chunk_start // ROWS_PER_CHUNK) + 1
            docs.append(LCDoc(
                page_content=page_text,
                metadata={
                    'source':       file_path,
                    'title':        getattr(document, 'title', os.path.basename(file_path)),
                    'page':         page_num,
                    'sheet':        sheet_name,
                    'content_type': 'spreadsheet',
                    'file_type':    ext.lstrip('.'),
                }
            ))
    return docs


# ── Core indexer ──────────────────────────────────────────────────────────────

def _index_in_background(document_id: int):
    """
    Index one document. Called exclusively by the single queue worker thread
    so there is never more than one indexing job running at a time.
    """
    try:
        from .models import Document, DocumentEmbedding
        from .rag_views import get_rag_chatbot

        document = Document.objects.get(id=document_id)
        if not document.file:
            return

        file_path = document.file.path
        ext = os.path.splitext(file_path)[1].lower()

        if ext not in SUPPORTED_EXTENSIONS:
            return

        # Ensure an embedding record exists
        DocumentEmbedding.objects.get_or_create(document=document)

        # Atomic claim — guards against duplicate queue entries for the same doc
        claimed = DocumentEmbedding.objects.filter(
            document=document,
            is_indexed=False,
            index_status__in=['pending', 'failed'],
        ).update(index_status='processing')

        if not claimed:
            # Already indexed or currently being processed
            logger.debug(f"[RAG] Doc {document_id} already claimed — skipping")
            return

        embedding = DocumentEmbedding.objects.get(document=document)
        logger.info(f"[RAG] Indexing started ({ext}): {document.title!r}")

        # get_rag_chatbot() blocks until initialization completes (lock-protected)
        chatbot = get_rag_chatbot()

        # Use org id if available, otherwise fall back to a user-scoped key
        org_id = (
            str(document.owner.organization_id)
            if document.owner.organization_id
            else f'user_{document.owner.id}'
        )

        # ── Route by file type ────────────────────────────────────────────────
        if ext == '.pdf':
            chatbot.index_documents(
                pdf_path=file_path,
                extract_tables=chatbot.config.ENABLE_TABLE_EXTRACTION,
                describe_images=chatbot.config.ENABLE_IMAGE_DESCRIPTION,
                org_id=org_id,
            )
            stats = chatbot.document_processor.get_processing_stats()
            chunk_count = stats.get('total_pages', 0)

        elif ext in {'.txt', '.md', '.text'}:
            lc_docs = _load_text_file(file_path, document)
            if not lc_docs:
                embedding.mark_failed("File appears to be empty")
                return
            chatbot.index_documents(documents=lc_docs, org_id=org_id)
            chunk_count = len(lc_docs)

        elif ext in {'.docx', '.doc'}:
            lc_docs = _load_docx_file(file_path, document)
            if not lc_docs:
                embedding.mark_failed("Could not extract text from .docx")
                return
            chatbot.index_documents(documents=lc_docs, org_id=org_id)
            chunk_count = len(lc_docs)

        elif ext in {'.xls', '.xlsx'}:
            lc_docs = _load_excel_file(file_path, document)
            if not lc_docs:
                embedding.mark_failed("Could not extract text from Excel file")
                return
            chatbot.index_documents(documents=lc_docs, org_id=org_id)
            chunk_count = len(lc_docs)

        else:
            embedding.mark_failed(f"Unsupported extension: {ext}")
            return

        embedding.mark_completed(
            chunk_count=chunk_count,
            embedding_model=chatbot.config.EMBEDDING_MODEL,
        )
        logger.info(f"[RAG] ✓ Indexed {document.title!r} — {chunk_count} chunks ({ext})")

    except Exception as exc:
        logger.error(
            f"[RAG] Indexing failed for document {document_id}: {exc}",
            exc_info=True,
        )
        try:
            from .models import DocumentEmbedding
            emb = DocumentEmbedding.objects.get(document_id=document_id)
            emb.mark_failed(str(exc))
        except Exception:
            pass
